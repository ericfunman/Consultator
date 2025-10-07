"""
Module de gestion de la liste des consultants
Fonctions pour afficher, filtrer et rechercher les consultants
"""

import os
import re
import sys
from datetime import datetime
from typing import List
from typing import Optional

import pandas as pd
import streamlit as st

# Constantes pour les colonnes du DataFrame
PRENOM_COL = "Pr√©nom"
NOM_COL = "Nom"
EMAIL_COL = "Email"
TELEPHONE_COL = "T√©l√©phone"
SALAIRE_COL = "Salaire annuel"
DISPONIBILITE_COL = "Disponibilit√©"
DATE_DISPONIBILITE_COL = "Date disponibilit√©"
GRADE_COL = "Grade"
TYPE_CONTRAT_COL = "Type contrat"
PRACTICE_COL = "Practice"
ENTITE_COL = "Entit√©"
DATE_CREATION_COL = "Date cr√©ation"
ID_COL = "ID"

# Constantes pour les statuts de disponibilit√©
STATUS_DISPONIBLE = "‚úÖ Disponible"
STATUS_EN_MISSION = "üî¥ En mission"

# Constantes pour les filtres
FILTRE_TOUS = "Tous"
FILTRE_DISPONIBLE = "Disponible"
FILTRE_EN_MISSION = "En mission"

# Constantes pour les messages
MSG_AUCUN_CONSULTANT = "‚ÑπÔ∏è Aucun consultant trouv√© dans la base de donn√©es"
MSG_AUCUN_RESULTAT = "‚ÑπÔ∏è Aucun consultant ne correspond aux crit√®res de recherche"
MSG_SERVICES_INDISPONIBLES = "‚ùå Les services de base ne sont pas disponibles"
MSG_ERREUR_CHARGEMENT = "‚ùå Erreur lors du chargement de la liste des consultants"
MSG_ERREUR_EXPORT = "‚ùå Erreur lors de l'export Excel"
MSG_ERREUR_RAPPORT = "‚ùå Erreur lors de la g√©n√©ration du rapport"
MSG_SUCCESS_EXPORT = "‚úÖ Fichier Excel g√©n√©r√© avec succ√®s !"
MSG_SUCCESS_RAPPORT = "‚úÖ Rapport g√©n√©r√© avec succ√®s !"
MSG_CONSULTANT_SELECTIONNE = "‚úÖ Consultant s√©lectionn√© : **{name}** (ID: {id})"

# Constantes pour les valeurs par d√©faut
DEFAULT_PRACTICE = "Non affect√©"
DEFAULT_DATE = "N/A"

# Ajouter les chemins n√©cessaires
current_dir = os.path.dirname(__file__)
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# Variables pour les imports
ConsultantService = None
get_database_session = None
Consultant = None
Competence = None
ConsultantCompetence = None
ConsultantSalaire = None
Mission = None
imports_ok = False

try:
    from sqlalchemy.orm import joinedload

    from database.database import get_database_session
    from database.models import Competence
    from database.models import Consultant
    from database.models import ConsultantCompetence
    from database.models import ConsultantSalaire
    from database.models import Mission
    from services.consultant_service import ConsultantService

    imports_ok = True
except ImportError:
    # Imports √©chou√©s, on continue quand m√™me
    pass


def _convert_consultants_to_dataframe(consultants: List) -> pd.DataFrame:
    """Convertit une liste de consultants en DataFrame"""
    consultants_data = []
    for consultant in consultants:
        practice_name = consultant.practice.nom if consultant.practice else DEFAULT_PRACTICE

        consultants_data.append(
            {
                ID_COL: consultant.id,
                PRENOM_COL: consultant.prenom,
                NOM_COL: consultant.nom,
                EMAIL_COL: consultant.email,
                TELEPHONE_COL: consultant.telephone,
                SALAIRE_COL: consultant.salaire_actuel or 0,
                DISPONIBILITE_COL: (STATUS_DISPONIBLE if consultant.disponibilite else STATUS_EN_MISSION),
                DATE_DISPONIBILITE_COL: consultant.date_disponibilite,
                GRADE_COL: consultant.grade,
                TYPE_CONTRAT_COL: consultant.type_contrat,
                PRACTICE_COL: practice_name,
                ENTITE_COL: consultant.entite or "N/A",
                DATE_CREATION_COL: (
                    consultant.date_creation.strftime("%d/%m/%Y") if consultant.date_creation else DEFAULT_DATE
                ),
            }
        )

    return pd.DataFrame(consultants_data)


def _create_search_filters(df: pd.DataFrame) -> tuple:
    """Cr√©e les widgets de recherche et de filtres"""
    col1, col2, col3, col4 = st.columns([2, 2, 1, 1])

    with col1:
        search_term = st.text_input(
            "üîç Recherche",
            placeholder="Nom, pr√©nom, email...",
            help="Rechercher par nom, pr√©nom ou email",
        )

    with col2:
        practice_filter = st.selectbox(
            "üè¢ Filtrer par practice",
            options=[FILTRE_TOUS] + sorted(set(df[PRACTICE_COL].tolist())),
            help="Filtrer les consultants par practice",
        )

    with col3:
        entite_filter = st.selectbox(
            "üèõÔ∏è Filtrer par entit√©",
            options=[FILTRE_TOUS] + sorted(set(df[ENTITE_COL].tolist())),
            help="Filtrer les consultants par entit√©",
        )

    with col4:
        availability_filter = st.selectbox(
            "üìä Statut",
            options=[FILTRE_TOUS, FILTRE_DISPONIBLE, FILTRE_EN_MISSION],
            help="Filtrer par disponibilit√©",
        )

    return search_term, practice_filter, entite_filter, availability_filter


def _apply_filters(
    df: pd.DataFrame,
    search_term: str,
    practice_filter: str,
    entite_filter: str,
    availability_filter: str,
) -> pd.DataFrame:
    """Applique les filtres au DataFrame"""
    filtered_df = df.copy()

    if search_term:
        filtered_df = filtered_df[
            filtered_df[PRENOM_COL].str.contains(search_term, case=False, na=False)
            | filtered_df[NOM_COL].str.contains(search_term, case=False, na=False)
            | filtered_df[EMAIL_COL].str.contains(search_term, case=False, na=False)
        ]

    if practice_filter != FILTRE_TOUS:
        filtered_df = filtered_df[filtered_df[PRACTICE_COL] == practice_filter]

    if entite_filter != FILTRE_TOUS:
        filtered_df = filtered_df[filtered_df[ENTITE_COL] == entite_filter]

    if availability_filter != FILTRE_TOUS:
        status_map = {
            FILTRE_DISPONIBLE: STATUS_DISPONIBLE,
            FILTRE_EN_MISSION: STATUS_EN_MISSION,
        }
        filtered_df = filtered_df[filtered_df[DISPONIBILITE_COL] == status_map[availability_filter]]

    return filtered_df


def _display_statistics(filtered_df: pd.DataFrame) -> None:
    """Affiche les statistiques des consultants"""
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        total_consultants = len(filtered_df)
        st.metric("üë• Total consultants", total_consultants)

    with col2:
        available_count = len(filtered_df[filtered_df[DISPONIBILITE_COL] == STATUS_DISPONIBLE])
        st.metric("‚úÖ Disponibles", available_count)

    with col3:
        busy_count = len(filtered_df[filtered_df[DISPONIBILITE_COL] == STATUS_EN_MISSION])
        st.metric("üî¥ En mission", busy_count)

    with col4:
        total_salary = filtered_df[SALAIRE_COL].sum()
        st.metric("üí∞ Masse salariale", f"{total_salary:,.0f}‚Ç¨")


def _get_display_columns() -> List[str]:
    """Retourne la liste des colonnes √† afficher dans le tableau"""
    return [
        PRENOM_COL,
        NOM_COL,
        EMAIL_COL,
        DISPONIBILITE_COL,
        DATE_DISPONIBILITE_COL,
        GRADE_COL,
        TYPE_CONTRAT_COL,
        PRACTICE_COL,
        ENTITE_COL,
    ]


def _create_column_config() -> dict:
    """Cr√©e la configuration des colonnes pour l'affichage du DataFrame"""
    return {
        PRENOM_COL: st.column_config.TextColumn(PRENOM_COL, width="small"),
        NOM_COL: st.column_config.TextColumn(NOM_COL, width="small"),
        EMAIL_COL: st.column_config.TextColumn(EMAIL_COL, width="large"),
        DISPONIBILITE_COL: st.column_config.TextColumn(DISPONIBILITE_COL, width="small"),
        DATE_DISPONIBILITE_COL: st.column_config.TextColumn(DATE_DISPONIBILITE_COL, width="small"),
        GRADE_COL: st.column_config.TextColumn(GRADE_COL, width="small"),
        TYPE_CONTRAT_COL: st.column_config.TextColumn(TYPE_CONTRAT_COL, width="small"),
        PRACTICE_COL: st.column_config.TextColumn(PRACTICE_COL, width="medium"),
        ENTITE_COL: st.column_config.TextColumn(ENTITE_COL, width="medium"),
    }


def _handle_consultant_selection(event, filtered_df: pd.DataFrame) -> None:
    """G√®re la s√©lection d'un consultant dans le tableau"""
    if event.selection.rows:
        selected_row = event.selection.rows[0]
        selected_consultant_data = filtered_df.iloc[selected_row]
        selected_id = int(selected_consultant_data[ID_COL])
        selected_name = f"{selected_consultant_data[PRENOM_COL]} {selected_consultant_data[NOM_COL]}"

        st.success(MSG_CONSULTANT_SELECTIONNE.format(name=selected_name, id=selected_id))

        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button(
                "üëÅÔ∏è Voir le profil",
                type="primary",
                width="stretch",
                key=f"view_profile_{selected_id}",
            ):
                st.session_state.view_consultant_profile = selected_id
                st.rerun()

        with col2:
            if st.button(
                "‚úèÔ∏è Modifier",
                width="stretch",
                key=f"edit_profile_{selected_id}",
            ):
                st.session_state.view_consultant_profile = selected_id
                st.session_state.edit_consultant_mode = True
                st.rerun()

        with col3:
            if st.button(
                "üóëÔ∏è Supprimer",
                width="stretch",
                key=f"delete_profile_{selected_id}",
            ):
                st.session_state.view_consultant_profile = selected_id
                st.session_state.delete_consultant_mode = True
                st.rerun()


def _handle_alternative_selection(filtered_df: pd.DataFrame) -> None:
    """G√®re la s√©lection alternative avec selectbox"""
    st.markdown("---")
    st.markdown("### üîç S√©lection alternative")

    selected_consultant = st.selectbox(
        "üë§ Ou s√©lectionner un consultant pour voir son profil d√©taill√©",
        options=[""] + [f"{row[PRENOM_COL]} {row[NOM_COL]} (ID: {row[ID_COL]})" for _, row in filtered_df.iterrows()],
        help="Choisissez un consultant pour acc√©der √† son profil complet",
    )

    if selected_consultant:
        match = re.search(r"\(ID: (\d+)\)", selected_consultant)
        if match:
            consultant_id = int(match.group(1))

            if st.button(
                "üëÅÔ∏è Voir le profil d√©taill√©",
                key=f"view_profile_alt_{consultant_id}",
            ):
                st.session_state.view_consultant_profile = consultant_id
                st.rerun()


def _display_action_buttons(filtered_df: pd.DataFrame) -> None:
    """Affiche les boutons d'actions group√©es"""
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("üìä Exporter en Excel", help="T√©l√©charger la liste au format Excel"):
            export_to_excel(filtered_df)

    with col2:
        if st.button("üìà G√©n√©rer rapport", help="Cr√©er un rapport d√©taill√©"):
            generate_consultants_report(filtered_df)

    with col3:
        if st.button("üîÑ Actualiser", help="Recharger les donn√©es"):
            st.rerun()


def show_consultants_list():
    """Affiche la liste des consultants avec filtres et recherche"""

    if not imports_ok:
        st.error(MSG_SERVICES_INDISPONIBLES)
        return

    st.markdown("### üìã Liste des consultants")

    # R√©cup√©rer tous les consultants
    try:
        with get_database_session() as session:
            consultants = session.query(Consultant).options(joinedload(Consultant.practice)).all()

        if not consultants:
            st.info(MSG_AUCUN_CONSULTANT)
            return

        # Convertir en DataFrame pour faciliter la manipulation
        df = _convert_consultants_to_dataframe(consultants)

        # Filtres et recherche
        search_term, practice_filter, entite_filter, availability_filter = _create_search_filters(df)

        # Appliquer les filtres
        filtered_df = _apply_filters(df, search_term, practice_filter, entite_filter, availability_filter)

        # Statistiques
        _display_statistics(filtered_df)

        st.markdown("---")

        # Affichage du tableau
        if filtered_df.empty:
            st.info(MSG_AUCUN_RESULTAT)
        else:
            # Configuration des colonnes √† afficher
            display_columns = _get_display_columns()

            # Afficher le DataFrame avec s√©lection interactive
            event = st.dataframe(
                filtered_df[display_columns],
                width="stretch",
                hide_index=True,
                on_select="rerun",
                selection_mode="single-row",
                column_config=_create_column_config(),
            )

            # Actions sur s√©lection
            _handle_consultant_selection(event, filtered_df)

            # S√©lection alternative avec selectbox
            _handle_alternative_selection(filtered_df)

            # Boutons d'actions group√©es
            _display_action_buttons(filtered_df)

    except Exception as e:
        st.error(f"{MSG_ERREUR_CHARGEMENT}: {e}")
        st.code(str(e))


def export_to_excel(df: pd.DataFrame):
    """Exporte les donn√©es des consultants vers Excel"""

    try:
        # Cr√©er un buffer pour le fichier Excel
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.styles import PatternFill

        # Cr√©er le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultants"

        # En-t√™tes avec style
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")

        # Donn√©es
        for row_num, row in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)  # Largeur max de 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Bouton de t√©l√©chargement
        st.download_button(
            label="üì• T√©l√©charger Excel",
            data=buffer,
            file_name=f"consultants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.success(MSG_SUCCESS_EXPORT)

    except ImportError:
        st.error("‚ùå Module openpyxl non install√©. Installez-le avec : pip install openpyxl")
    except Exception as e:
        st.error(f"{MSG_ERREUR_EXPORT}: {e}")


def generate_consultants_report(df: pd.DataFrame):
    """G√©n√®re un rapport d√©taill√© des consultants"""

    try:
        st.markdown("### üìä Rapport des consultants")

        # Statistiques g√©n√©rales
        col1, col2, col3 = st.columns(3)

        with col1:
            st.subheader("üë• Effectif")
            st.metric("Total", len(df))
            available = len(df[df[DISPONIBILITE_COL] == STATUS_DISPONIBLE])
            st.metric("Disponibles", f"{available} ({available / len(df) * 100:.1f}%)")

        with col2:
            st.subheader("üí∞ R√©mun√©ration")
            total_salary = df[SALAIRE_COL].sum()
            avg_salary = df[SALAIRE_COL].mean()
            st.metric("Masse salariale", f"{total_salary:,.0f}‚Ç¨")
            st.metric("Salaire moyen", f"{avg_salary:,.0f}‚Ç¨")

        with col3:
            st.subheader("üè¢ R√©partition")
            practice_counts = df[PRACTICE_COL].value_counts()
            for practice, count in practice_counts.items():
                st.metric(practice, count)

        # Graphiques
        st.markdown("### üìà Visualisations")

        # R√©partition par practice
        if len(practice_counts) > 1:
            st.bar_chart(practice_counts)

        # Distribution des salaires
        if len(df) > 1:
            salary_data = df[[PRENOM_COL, NOM_COL, SALAIRE_COL]].copy()
            salary_data["Nom complet"] = salary_data[PRENOM_COL] + " " + salary_data[NOM_COL]
            salary_data = salary_data.sort_values(SALAIRE_COL, ascending=False)

            st.bar_chart(
                salary_data.set_index("Nom complet")[SALAIRE_COL],
                width="stretch",
            )

        st.success(MSG_SUCCESS_RAPPORT)

    except Exception as e:
        st.error(f"{MSG_ERREUR_RAPPORT}: {e}")


def show_consultants_list_table(consultants=None, filters=None):  # noqa: ARG001
    """
    Affiche la liste des consultants sous forme de tableau
    Alias de show_consultants_list pour compatibilit√© avec les tests
    
    Args:
        consultants: Liste de consultants (optionnel, recharg√© si None)
        filters: Filtres √† appliquer (optionnel)
    """
    # Appeler la fonction principale
    show_consultants_list()

